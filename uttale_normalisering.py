import os
import re
import openpyxl

basedir = os.path.dirname(__file__)
path = os.path.join(basedir, '1800 - Nordland 14.06.2023, redigert av AiN .xlsx')

import openpyxl
import shutil
import os

base, ext = os.path.splitext(path)
copy_file_path = f"{base}_copy{ext}"
shutil.copy2(path, copy_file_path)
print(f"Copied file to: {copy_file_path}")

# Load the copied workbook
wb = openpyxl.load_workbook(copy_file_path)
ws = wb.active

col_idx = 11  # Column 11 = 'K'

# Iterate over the cells in column 11, skipping the header
for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
    cell = row[0]
    if cell.value:
        fixed = cell.value
        fixed = re.sub(r'[=@*<>_]', '', fixed)
        fixed = fixed.strip()

        fixed = fixed.replace('`',"'")
        fixed = fixed.replace('²','ɳ')
        fixed = fixed.replace('±', 'ŋ')
        fixed = fixed.replace('¼', 'ʈ')
        fixed = fixed.replace('ª', 'ɽ')
        fixed = fixed.replace('¿', 'ʃ')
        fixed = fixed.replace('¾', 'ʂ')
        fixed = fixed.replace('®', 'ç')
        fixed = fixed.replace('³', 'ɲ')
        fixed = fixed.replace('¬', 'ʎ')
        fixed = fixed.replace('½', 'c')
        fixed = fixed.replace('«', 'ɭ')
        if len(fixed) > 1:
            fixed = fixed.replace('?', "'")


        if(fixed[0]) == "'":
            fixed = fixed.replace("'", '¹', 1)

        #Disse er til hvis det er flere alternative uttaler. Funker sånn nesten helt. Problem hvis ' er på slutten, men er sluttparantes bak f.eks.
        #Også inconsistent med mellomrom format.
        # fixed = re.sub(r" '(?=\S)", " ¹", fixed)
        # fixed = re.sub(r",\'(?=\S)", " ¹", fixed)
        # fixed = re.sub(r";\'(?=\S)", " ¹", fixed)
        # fixed = re.sub(r"¤\'(?=\S)", " ¹", fixed)

            
        fixed = fixed.replace('"', '²')

        ws.cell(row=row[0].row, column=col_idx).value = fixed


#Finner alle tegn brukt i uttale-kolonnen

def unique_chars(path):

    wb = openpyxl.load_workbook(copy_file_path)
    ws = wb.active

    unique_chars = set()

    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
        cell = row[0]
        if cell.value is not None:
            unique_chars.update(str(cell.value))

    # Convert the set to a sorted string (optional sorting)
    print(''.join(sorted(unique_chars)))



# Save the modified copy
wb.save(copy_file_path)
print(f"Changes saved to: {copy_file_path}")
# unique_chars(copy_file_path)
